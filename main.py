from imbox import Imbox
from datetime import datetime, timedelta
import pandas as pd
from barcode_reader import BarcodeReader
from barcode_reader import digitable_line
from openpyxl import Workbook
import os

username = '[YOUR_EMAIL]'
password = open('passwords/token', 'r').read()
host = "imap.gmail.com"
download_folder = "[TICKET_FOLDER_NAME]"

mail = Imbox(host, username=username, password=password, ssl=True)
messages = mail.messages(date__gt=datetime.today() - timedelta(days=20), raw='has:attachment')

#for pass_ in ['123']:
#    while convert_from_path('arquivo.pdf', userpw=pass_)

wb = Workbook()
ws = wb.active
r = 1
ws.cell(row=1, column=1).value = 'Subject'
ws.cell(row=1, column=2).value = 'Barcode'
ws.cell(row=1, column=3).value = 'Digitable line'
ws.cell(row=1, column=4).value = 'Filename'

for (uid, message) in messages:
    if len(message.attachments) > 0:
        for attach in message.attachments:
            att_file = attach["filename"]

            if '.pdf' in att_file:
                download_path = f'{download_folder}\\{att_file}'

                with open(download_path, 'wb') as fp:
                    fp.write(attach['content'].read())

                try:
                    barcode = BarcodeReader(download_path)
                    linha_dig = digitable_line(barcode)
                except:
                    barcode = False

                if not barcode:
                   os.remove(download_path)
                else:
                    print(message.subject, '-', barcode)
                    r += 1
                    ws.cell(row=r, column=1).value = str(message.subject)
                    ws.cell(row=r, column=2).value = barcode
                    ws.cell(row=r, column=3).value = linha_dig
                    ws.cell(row=r, column=4).value = att_file

wb.save('tickets.xlsx')
mail.logout()